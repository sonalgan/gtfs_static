import argparse
from openpyxl import load_workbook
import pandas as pd
import os
import warnings

def extract_timetables(file_path):
    """
    Extracts timetables from the given Excel file.

    Args:
        file_path (str): Path to the Excel file.

    Returns:
        list: List of timetables, where each timetable is a list of rows.
    """
    workbook = load_workbook(filename=file_path, read_only=True)
    worksheet = workbook.active

    start_marker = "BENGALURU METROPOLITAN TRANSPORT CORPORATION" 
    timetables = []
    current_timetable = []

    for row in worksheet.iter_rows(values_only=True):
        if row[0] and str(row[0]).strip() == start_marker:
            if current_timetable:
                timetables.append(current_timetable)
            current_timetable = [row]
        elif current_timetable:
            current_timetable.append(row)

    if current_timetable:
        timetables.append(current_timetable)

    return timetables

def extract_route_info(timetable):
    """
    Extracts route information from the given timetable.

    Args:
        timetable (list): List of rows representing a timetable.

    Returns:
        dict: Route information.
    """
    route_info = {"Route_Code": None, "Route_Origin": None, "Route_Destination": None}
    for row in timetable:
        if isinstance(row[0], str):
            if row[0].strip().lower().startswith("route"):
                route_info["Route_Code"] = row[1]
            elif row[0].strip().lower() in ["brand", "vajra"]:
                if row[2]:
                    route_info["Route_Origin"] = row[2] if route_info["Route_Origin"] is None else route_info["Route_Origin"]
                    route_info["Route_Destination"] = row[2] if route_info["Route_Destination"] is None else route_info["Route_Destination"]
    return route_info

def convert_excel_to_gtfs(file_path, output_folder):
    """
    Converts BMTC Bus Schedule Excel file to GTFS format.

    Args:
        file_path (str): Path to the Excel file.
        output_folder (str): Path to the output folder for GTFS files.
    """
    timetables = extract_timetables(file_path)
    routes = []

    for t_id, t in enumerate(timetables):
        route_info = extract_route_info(t)
        route_info['timetable_id'] = t_id
        routes.append(route_info)

    routes_df = pd.DataFrame(routes).drop_duplicates().reset_index(drop=True)

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        routes_df['route_id'] = routes_df.groupby('Route_Code').ngroup().add(1).apply(lambda x: f'route_{x}')
        routes_df['Route_Origin'] = routes_df['Route_Origin'].str.strip().str.replace(r'\s+TO$', '')
        routes_df['Route_Destination'] = routes_df['Route_Destination'].str.strip().str.replace(r'^TO\s+', '')

    schedules_df = routes_df[['route_id', 'timetable_id']].rename(columns={'timetable_id': 'schedule_id'})
    schedules_df['schedule_id'] = schedules_df['schedule_id'].apply(lambda x: f'schedule_{x}')

    routes_df[['route_id', 'Route_Code', 'Route_Origin', 'Route_Destination']].to_csv(os.path.join(output_folder, "routes_info.csv"), index=False)
    schedules_df.to_csv(os.path.join(output_folder, "schedules_info.csv"), index=False)

def main():
    parser = argparse.ArgumentParser(description='Convert BMTC Bus Schedule Excel file to GTFS format.')
    parser.add_argument('file_path', type=str, help='Path to the BMTC Bus Schedule Excel file')
    parser.add_argument('output_folder', type=str, help='Path to the output folder for GTFS files')
    args = parser.parse_args()

    file_path = args.file_path
    output_folder = args.output_folder

    convert_excel_to_gtfs(file_path, output_folder)

if __name__ == "__main__":
    main()
